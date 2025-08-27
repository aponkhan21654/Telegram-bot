# -*- coding: utf-8 -*-
import os
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters
from openpyxl import load_workbook

# ===== CONFIG =====
BOT_TOKEN = os.environ.get("BOT_TOKEN")  # Bot token now from environment variable
ADMIN_ID = 7867235273  # Admin Telegram ID

# ===== VARIABLES =====
ADMIN_FILES = {
    "0f Number": None,
    "Webmail": None,
    "Instagram ID": None
}
PRICES = {
    "0f Number": 0.0,
    "Webmail": 0.0,
    "Instagram ID": 0.0
}
TEMP = {}  # Temporary storage for actions

# ===== KEYBOARD =====
main_keyboard = ReplyKeyboardMarkup(
    [
        [KeyboardButton("0f Number"), KeyboardButton("Webmail")],
        [KeyboardButton("Instagram ID")],
        [KeyboardButton("Price Set"), KeyboardButton("Clear Admin File")]
    ],
    resize_keyboard=True
)

category_keyboard = ReplyKeyboardMarkup(
    [
        [KeyboardButton("0f Number"), KeyboardButton("Webmail")],
        [KeyboardButton("Instagram ID")],
        [KeyboardButton("Cancel")]
    ],
    resize_keyboard=True
)

# ===== START =====
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Welcome! Use the buttons below 👇", reply_markup=main_keyboard)

# ===== Admin File Upload =====
async def upload_admin_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global TEMP
    if update.message.from_user.id != ADMIN_ID:
        await update.message.reply_text("You are not authorized to upload admin file.")
        return

    if not update.message.document:
        await update.message.reply_text("Please send an Excel file.")
        return

    TEMP["action"] = "upload_file"
    TEMP["file"] = await update.message.document.get_file()
    await update.message.reply_text("Which category to save this file?", reply_markup=category_keyboard)

# ===== Admin Price Set =====
async def set_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global TEMP
    if update.message.from_user.id != ADMIN_ID:
        await update.message.reply_text("You are not authorized to set prices.")
        return
    TEMP["action"] = "set_price"
    await update.message.reply_text("Select category to set price:", reply_markup=category_keyboard)

# ===== Handle Category & Price Input =====
async def handle_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global TEMP
    text = update.message.text

    # ----- Price Set Handling -----
    if TEMP.get("action") == "waiting_price":
        category = TEMP["category"]
        try:
            price = float(text)
            PRICES[category] = price
            await update.message.reply_text(f"✅ Price for {category} set to {price} TK", reply_markup=main_keyboard)
        except:
            await update.message.reply_text("Invalid number. Enter number only.")
        TEMP.clear()
        return

    # ----- Admin File Upload -----
    if TEMP.get("action") == "upload_file" and text in ADMIN_FILES:
        file = TEMP["file"]
        file_path = f"admin_{text}.xlsx"
        await file.download_to_drive(file_path)
        ADMIN_FILES[text] = file_path
        await update.message.reply_text(f"Admin file saved for {text}.", reply_markup=main_keyboard)
        TEMP.clear()
        return

    # ----- Admin selects category to set price -----
    if TEMP.get("action") == "set_price" and text in ADMIN_FILES:
        TEMP["category"] = text
        TEMP["action"] = "waiting_price"
        await update.message.reply_text(f"Enter new price for {text} (number only):", reply_markup=ReplyKeyboardRemove())
        return

    # ----- User selects category to check file -----
    if text in ADMIN_FILES:
        if not ADMIN_FILES[text]:
            await update.message.reply_text(f"❌ Admin file for {text} not set yet.")
            return
        TEMP["action"] = "check_file"
        TEMP["category"] = text
        await update.message.reply_text(f"Please send your Excel file for {text}.\nPrice: {PRICES[text]} TK")
        return

# ===== Process User File =====
async def process_user_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global TEMP
    if TEMP.get("action") != "check_file":
        return
    category = TEMP["category"]

    file = await update.message.document.get_file()
    user_file_path = f"user_{update.message.from_user.id}_{category}.xlsx"
    await file.download_to_drive(user_file_path)

    # Load admin file
    wb_admin = load_workbook(ADMIN_FILES[category])
    ws_admin = wb_admin.active
    admin_numbers = {
        str(ws_admin[f"A{i}"].value).strip()
        for i in range(1, ws_admin.max_row + 1)
        if ws_admin[f"A{i}"].value not in (None, "")
    }

    # Load user file
    wb_user = load_workbook(user_file_path)
    ws_user = wb_user.active
    user_numbers = {
        str(ws_user[f"A{i}"].value).strip()
        for i in range(1, ws_user.max_row + 1)
        if ws_user[f"A{i}"].value not in (None, "")
    }

    matched_numbers = admin_numbers & user_numbers
    matched_count = len(matched_numbers)
    unmatched_count = len(user_numbers - admin_numbers)
    total_price = matched_count * PRICES[category]

    await update.message.reply_text(
        f"YOUR REPORT: {matched_count} OK✅\n"
        f"PRICE: {total_price} TK\n"
        f"BACK 👉 : {unmatched_count} BACK❌\n"
        f"Forward -  @arafatislamapon"
    )
    TEMP.clear()

# ===== Clear Admin Files =====
async def clear_admin_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global ADMIN_FILES
    if update.message.from_user.id != ADMIN_ID:
        await update.message.reply_text("You are not authorized.")
        return
    ADMIN_FILES = {key: None for key in ADMIN_FILES}
    await update.message.reply_text("All admin files cleared ✅", reply_markup=main_keyboard)

# ===== MAIN =====
if __name__ == "__main__":
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    # Handlers
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.Document.ALL & filters.User(ADMIN_ID), upload_admin_file))
    app.add_handler(MessageHandler(filters.Regex("^Price Set$"), set_price))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_category))
    app.add_handler(MessageHandler(filters.Document.ALL & ~filters.User(ADMIN_ID), process_user_file))
    app.add_handler(MessageHandler(filters.Regex("^Clear Admin File$"), clear_admin_file))

    print("Bot is running...")
    app.run_polling()
